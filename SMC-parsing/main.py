# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter, column_index_from_string
from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
from config import url

class Product:
    def __init__(self):
        self.id = 0
        self.article = ''
        self.prise = 0
        self.series = ''
        self.name = ''

    def constructor(self, a, b, s, n):
        self.id = a
        self.article = b
        self.series = s
        self.name = n


book_read = openpyxl.load_workbook("smc-test.xlsx")
a = book_read.sheetnames[0]
sheet_read = book_read[a]
massOfArt = []
# print(massOfArt)
i, j = 2, 1
while(sheet_read.cell(i, j).value != None):
    pr = Product()
    pr.constructor(sheet_read.cell(i, j).value,sheet_read.cell(i, j+1).value,sheet_read.cell(i, j+3).value,sheet_read.cell(i, j+4).value)
    massOfArt.append(pr)
    i +=1
print(i)
book_write_try = openpyxl.load_workbook("smc-test-rez.xlsx")
book_write_els = openpyxl.load_workbook("smc-test-error.xlsx")

sheet_write_try = book_write_try[book_write_try.sheetnames[0]]
sheet_write_els = book_write_els[book_write_els.sheetnames[0]]

sheet_write_try['A1'] = 'product_id'
sheet_write_try['B1'] = 'model'
sheet_write_try['C1'] = 'sku'
sheet_write_try['D1'] = 'price'


sheet_write_els['A1'] = 'product_id'
sheet_write_els['B1'] = 'model'
sheet_write_els['C1'] = 'sku'
sheet_write_els['D1'] = 'price'

sr_try, st_try = 2, 1
sr_els, st_els = 2, 1

for pr in massOfArt:
    try:
        browser = webdriver.Chrome()
        url_s = url + pr.article + ".html"
        browser.get(url_s)
        rez_parse = browser.find_element(by=By.ID, value="price").text
        sheet_write_try['A' + str(sr_try)] = pr.id
        sheet_write_try['B' + str(sr_try)] = pr.article
        sheet_write_try['C' + str(sr_try)] = pr.article
        sheet_write_try['D' + str(sr_try)] = rez_parse
        sr_try += 1
    except:
        sheet_write_els['A' + str(sr_els)] = pr.id
        sheet_write_els['B' + str(sr_els)] = pr.article
        sheet_write_els['C' + str(sr_els)] = pr.article
        sheet_write_els['D' + str(sr_els)] = 'Ошибка'
        sr_els += 1

print("Правильных: ", sr_try-1)
print("Неправильных: ", sr_els-1)

book_write_try.save("smc-test-rez.xlsx")
book_write_els.save("smc-test-error.xlsx")

#f = open("log.txt", "w")




#    url_s = 'https://www.smcpneumatics.com/MY1M50TN-100.html'
#    rez_parse = browser.find_element(by="id", value='price').text
#    rez_parse = browser.find_element(webdriver.common.by.By.ID, "price").text
    #current-search-title

#     try:
#         browser = webdriver.Chrome()
#         url_s = url + pr.article + ".html"
#         browser.get(url_s)
#         rez_parse = browser.find_element(by="id", value='price').text
#
#         sheet_write_try['A' + str(sr_try)] = pr.id
#         sheet_write_try['B' + str(sr_try)] = pr.article
#         sheet_write_try['C' + str(sr_try)] = pr.article
#         sheet_write_try['D' + str(sr_try)] = rez_parse
#         sr_try += 1
#     except:
#         sheet_write_els['A' + str(sr_els)] = pr.id
#         sheet_write_els['B' + str(sr_els)] = pr.article
#         sheet_write_els['C' + str(sr_els)] = pr.article
#         sheet_write_els['D' + str(sr_els)] = 'Ошибка'
#         sr_els += 1
#
#
# print("Правильных: ", sr_try-1)
# print("Неправильных: ", sr_els-1)
#
# book_write_try.save("smc-test-rez.xlsx")
# book_write_els.save("smc-test-error.xlsx")





#-------------------------------------------------------
# for pr in massOfArt:
#     response = requests.get(url + pr.article)
#     bs = BeautifulSoup(response.text, "lxml")
#     val = bs.find('div', {'itemprop':'price'})
#     print(val.text)
#-------------------------------------------------------
# art = 'MY1M50TN-100'
# # response = requests.get(url + art+'.html')
# response = requests.get('https://www.smcpneumatics.com/assets/templates/common-html5/js/jquery.simplemodal.min.js?_=1679612116416')
# #response.raise_for_status()
# print(response.status_code)
# # bs = BeautifulSoup(response.text, "html.parser")
# print(bs)
# val = bs.find('div', {'itemprop':'price'})
# print(val.text)











