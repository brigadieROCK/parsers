from selenium import webdriver
import openpyxl
from config import url
import time
#https://www.smcpneumatics.com/New-Search_ep_2254-1.html#query=MSQ&view=row&showed=100&page=0


class Product:
    def __init__(self):
        self.id = 0
        self.article = ''
        self.prise = 0
        self.series = ''
        self.name = ''

    def constructor(self, a, b, s, n):
        self.id = a
        self.article = b
        self.series = s
        self.name = n


# book_read = openpyxl.load_workbook("smc-test.xlsx")
# a = book_read.sheetnames[0]
# sheet_read = book_read[a]
# massOfArtInBook = []
# i, j = 2, 1
# while(sheet_read.cell(i, j).value != None):
#     pr = Product()
#     pr.constructor(sheet_read.cell(i, j).value,sheet_read.cell(i, j+1).value,sheet_read.cell(i, j+3).value,sheet_read.cell(i, j+4).value)
#     massOfArtInBook.append(pr)
#     i +=1



browser = webdriver.Chrome()
url_s = 'https://www.smcpneumatics.com/New-Search_ep_2254-1.html#query=MSQ&view=row&showed=100&page=0'
browser.get(url_s)
time.sleep(60)
rez_parse = browser.find_elements(by="class name", value='product-item alternative')
print(rez_parse)


